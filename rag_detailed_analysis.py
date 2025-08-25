import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from typing import List, Tuple, Dict
import os
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
import json
import re

class RAGDetailedAnalysis:
    """詳細なRAG検索結果の分析クラス"""
    
    def __init__(self, excel_file_path: str):
        """
        RAG詳細分析クラスの初期化
        
        Parameters:
        -----------
        excel_file_path : str
            分析対象のExcelファイルパス
        """
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.worksheet = None
        self.df = None
        self.analysis_results = None
        
    def load_excel(self, sheet_name: str = None):
        """Excelファイルを読み込み"""
        try:
            # pandasでデータを読み込み
            excel_data = pd.read_excel(self.excel_file_path, sheet_name=sheet_name)
            
            # 複数シートの場合は辞書形式で返されるため、適切に処理
            if isinstance(excel_data, dict):
                print(f"複数シートが検出されました。利用可能なシート: {list(excel_data.keys())}")
                if sheet_name is None:
                    # シート名が指定されていない場合は最初のシートを使用
                    sheet_name = list(excel_data.keys())[0]
                    print(f"最初のシート '{sheet_name}' を使用します。")
                elif sheet_name not in excel_data:
                    print(f"指定されたシート '{sheet_name}' が見つかりません。最初のシートを使用します。")
                    sheet_name = list(excel_data.keys())[0]
                self.df = excel_data[sheet_name]
            else:
                self.df = excel_data
            
            self.workbook = openpyxl.load_workbook(self.excel_file_path)
            
            # シート名の処理
            if sheet_name is None:
                self.worksheet = self.workbook.active
            else:
                try:
                    self.worksheet = self.workbook[sheet_name]
                except KeyError:
                    print(f"警告: シート '{sheet_name}' が見つかりません。最初のシートを使用します。")
                    self.worksheet = self.workbook.active
            
            print(f"Excelファイルを読み込みました: {self.excel_file_path}")
            print(f"使用シート: {sheet_name if sheet_name else '最初のシート'}")
            print(f"データ形状: {self.df.shape}")
            
        except Exception as e:
            print(f"Excelファイルの読み込みに失敗しました: {e}")
            raise
    
    def is_colored_cell(self, cell) -> bool:
        """セルが塗りつぶされているかチェック（白または無色以外）"""
        # 複数の方法で色情報を取得
        fill_objects = []
        
        # 方法1: 直接のfill属性
        if hasattr(cell, 'fill') and cell.fill:
            fill_objects.append(cell.fill)
        
        # 方法2: style.fill属性
        if hasattr(cell, 'style') and hasattr(cell.style, 'fill') and cell.style.fill:
            fill_objects.append(cell.style.fill)
        
        # 方法3: StyleProxyの内部オブジェクト
        for fill in fill_objects[:]:  # コピーを作成してループ中に変更
            if hasattr(fill, '_style') and hasattr(fill._style, 'fill'):
                fill_objects.append(fill._style.fill)
            elif hasattr(fill, 'fill'):
                fill_objects.append(fill.fill)
        
        # 各fillオブジェクトをチェック
        for fill in fill_objects:
            if not fill:
                continue
                
            if not isinstance(fill, PatternFill):
                continue
                
            fg_color = fill.fgColor
            bg_color = fill.bgColor
            pattern_type = fill.patternType
            
            # パターンが設定されている場合
            if pattern_type and pattern_type != 'none':
                return True
            
            # 前景色の判定
            if fg_color and fg_color.type != "auto":
                if self._is_colored_color(fg_color):
                    return True
            
            # 背景色の判定
            if bg_color and bg_color.type != "auto":
                if self._is_colored_color(bg_color):
                    return True
                    
        return False
    
    def _is_colored_color(self, color) -> bool:
        """Colorオブジェクトが白以外の色かどうかを判定"""
        if not color:
            return False
            
        if color.type == "rgb":
            # RGB値で白を判定
            if color.rgb and color.rgb.upper() in ["FFFFFF", "FFFFFFFF", "00000000"]:
                return False
            # RGB値が設定されている場合は塗りつぶされている
            if color.rgb:
                return True
        elif color.type == "theme":
            # テーマ色で白を判定（一般的な白のテーマ番号）
            if color.theme in [0, 1]:  # 白系のテーマ番号
                return False
            # テーマ色が設定されている場合は塗りつぶされている
            if color.theme is not None:
                return True
        elif color.type == "indexed":
            # インデックス色で白を判定
            if color.indexed in [0, 1]:  # 白系のインデックス番号
                return False
            # インデックス色が設定されている場合は塗りつぶされている
            if color.indexed is not None:
                return True
                
        return False
    
    def debug_cell_color(self, cell) -> str:
        """セルの色情報をデバッグ用に取得"""
        # 複数の方法で色情報を取得
        fill_objects = []
        
        # 方法1: 直接のfill属性
        if hasattr(cell, 'fill') and cell.fill:
            fill_objects.append(cell.fill)
        
        # 方法2: style.fill属性
        if hasattr(cell, 'style') and hasattr(cell.style, 'fill') and cell.style.fill:
            fill_objects.append(cell.style.fill)
        
        # 方法3: StyleProxyの内部オブジェクト
        for fill in fill_objects[:]:  # コピーを作成してループ中に変更
            if hasattr(fill, '_style') and hasattr(fill._style, 'fill'):
                fill_objects.append(fill._style.fill)
            elif hasattr(fill, 'fill'):
                fill_objects.append(fill.fill)
        
        if not fill_objects:
            return "No fill objects found"
        
        color_info = []
        for i, fill in enumerate(fill_objects):
            if not fill:
                color_info.append(f"Fill{i}: None")
                continue
                
            if not isinstance(fill, PatternFill):
                color_info.append(f"Fill{i}: {type(fill)}")
                continue
                
            fg_color = fill.fgColor
            bg_color = fill.bgColor
            pattern_type = fill.patternType
            
            fill_info = f"Fill{i}: Pattern={pattern_type}"
            
            # 前景色の詳細情報
            if fg_color:
                fill_info += f", FG={self._get_color_info(fg_color)}"
            else:
                fill_info += ", FG=None"
                
            # 背景色の詳細情報
            if bg_color:
                fill_info += f", BG={self._get_color_info(bg_color)}"
            else:
                fill_info += ", BG=None"
                
            color_info.append(fill_info)
            
        return " | ".join(color_info)
    
    def _get_color_info(self, color) -> str:
        """Colorオブジェクトから色情報を取得"""
        if not color:
            return "None"
            
        color_info = f"{color.type}"
        
        if color.type == "rgb":
            color_info += f"={color.rgb}"
        elif color.type == "theme":
            color_info += f"={color.theme}"
        elif color.type == "indexed":
            color_info += f"={color.indexed}"
        elif color.type == "auto":
            color_info += "=auto"
        else:
            color_info += f"=unknown({color})"
            
        return color_info
    
    def get_rag_columns(self) -> List[str]:
        """RAGカラム（RAG1〜RAG10）を取得"""
        rag_columns = []
        for i in range(1, 11):
            column_name = f"RAG{i}"
            if column_name in self.df.columns:
                rag_columns.append(column_name)
        return rag_columns
    
    def extract_score_and_similarity(self, text: str) -> Dict[str, float]:
        """
        RAGカラムのテキストからScoreと類似度を抽出
        
        Parameters:
        -----------
        text : str
            抽出対象のテキスト
            
        Returns:
        --------
        Dict[str, float]
            Scoreと類似度の値（抽出できない場合はNone）
        """
        if pd.isna(text) or not isinstance(text, str):
            return {'score': None, 'similarity': None}
        
        result = {'score': None, 'similarity': None}
        
        # Scoreの抽出パターン
        score_patterns = [
            r'Score[:\s]*([0-9]+\.?[0-9]*)',
            r'スコア[:\s]*([0-9]+\.?[0-9]*)',
            r'検索スコア[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*\(Score\)',
            r'Score\s*=\s*([0-9]+\.?[0-9]*)'
        ]
        
        # 類似度の抽出パターン
        similarity_patterns = [
            r'類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'LLM回答\s*⇔\s*【回答】類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'回答類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*\(類似度\)',
            r'Similarity[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*%?\s*\(類似度\)'
        ]
        
        # Scoreの抽出
        for pattern in score_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    result['score'] = float(match.group(1))
                    break
                except ValueError:
                    continue
        
        # 類似度の抽出
        for pattern in similarity_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    result['similarity'] = float(match.group(1))
                    break
                except ValueError:
                    continue
        
        return result
    
    def analyze_rag_adoption_detailed(self, debug_colors: bool = False) -> pd.DataFrame:
        """
        詳細なRAG採択分析を実行
        
        Parameters:
        -----------
        debug_colors : bool, optional
            色情報のデバッグ出力を行うかどうか（デフォルト: False）
        """
        if self.df is None:
            raise ValueError("Excelファイルが読み込まれていません。")
        
        rag_columns = self.get_rag_columns()
        if not rag_columns:
            raise ValueError("RAGカラム（RAG1〜RAG10）が見つかりません。")
        
        print(f"分析対象RAGカラム: {rag_columns}")
        
        results = []
        adoption_patterns = []
        
        # 全体のScoreと類似度の統計用
        all_scores = []
        all_similarities = []
        adopted_scores = []
        adopted_similarities = []
        
        for row_idx, row_data in self.df.iterrows():
            adopted_count = 0
            adopted_rags = []
            adoption_pattern = []
            row_scores = []
            row_similarities = []
            adopted_row_scores = []
            adopted_row_similarities = []
            
            for rag_col in rag_columns:
                excel_row = row_idx + 2
                excel_col = self.df.columns.get_loc(rag_col) + 1
                cell = self.worksheet.cell(row=excel_row, column=excel_col)
                
                is_adopted = self.is_colored_cell(cell)
                adoption_pattern.append(1 if is_adopted else 0)
                
                # デバッグ出力（必要に応じて）
                if debug_colors and row_idx < 3:  # 最初の3行のみデバッグ出力
                    color_info = self.debug_cell_color(cell)
                    print(f"行{row_idx+1}, {rag_col}: {color_info} -> {'採択' if is_adopted else '未採択'}")
                
                # Scoreと類似度を抽出
                cell_text = str(row_data[rag_col]) if not pd.isna(row_data[rag_col]) else ""
                extracted_values = self.extract_score_and_similarity(cell_text)
                
                # 統計用に値を記録
                if extracted_values['score'] is not None:
                    row_scores.append(extracted_values['score'])
                    all_scores.append(extracted_values['score'])
                    if is_adopted:
                        adopted_row_scores.append(extracted_values['score'])
                        adopted_scores.append(extracted_values['score'])
                
                if extracted_values['similarity'] is not None:
                    row_similarities.append(extracted_values['similarity'])
                    all_similarities.append(extracted_values['similarity'])
                    if is_adopted:
                        adopted_row_similarities.append(extracted_values['similarity'])
                        adopted_similarities.append(extracted_values['similarity'])
                
                if is_adopted:
                    adopted_count += 1
                    adopted_rags.append(rag_col)
            
            adoption_rate = (adoption_count / len(rag_columns)) * 100 if rag_columns else 0
            
            # 採択パターンを記録
            pattern_str = ''.join(map(str, adoption_pattern))
            adoption_patterns.append(pattern_str)
            
            # 行内のScoreと類似度の統計
            row_score_stats = self._calculate_stats(row_scores)
            row_similarity_stats = self._calculate_stats(row_similarities)
            adopted_score_stats = self._calculate_stats(adopted_row_scores)
            adopted_similarity_stats = self._calculate_stats(adopted_row_similarities)
            
            result = {
                '行番号': row_idx + 1,
                '採択数': adopted_count,
                '総検索結果数': len(rag_columns),
                '採択率(%)': round(adoption_rate, 2),
                '採択されたRAG': ', '.join(adopted_rags) if adopted_rags else 'なし',
                '採択パターン': pattern_str,
                '採択パターン詳細': self._get_pattern_description(adoption_pattern, rag_columns),
                # Score統計
                'Score平均': row_score_stats['mean'],
                'Score最大': row_score_stats['max'],
                'Score最小': row_score_stats['min'],
                'Score標準偏差': row_score_stats['std'],
                '採択Score平均': adopted_score_stats['mean'],
                '採択Score最大': adopted_score_stats['max'],
                '採択Score最小': adopted_score_stats['min'],
                # 類似度統計
                '類似度平均': row_similarity_stats['mean'],
                '類似度最大': row_similarity_stats['max'],
                '類似度最小': row_similarity_stats['min'],
                '類似度標準偏差': row_similarity_stats['std'],
                '採択類似度平均': adopted_similarity_stats['mean'],
                '採択類似度最大': adopted_similarity_stats['max'],
                '採択類似度最小': adopted_similarity_stats['min']
            }
            
            # 元のデータも含める
            for col in self.df.columns:
                result[col] = row_data[col]
            
            results.append(result)
        
        # 全体統計を計算
        overall_score_stats = self._calculate_stats(all_scores)
        overall_similarity_stats = self._calculate_stats(all_similarities)
        overall_adopted_score_stats = self._calculate_stats(adopted_scores)
        overall_adopted_similarity_stats = self._calculate_stats(adopted_similarities)
        
        # 全体統計を保存
        self.overall_stats = {
            'score': overall_score_stats,
            'similarity': overall_similarity_stats,
            'adopted_score': overall_adopted_score_stats,
            'adopted_similarity': overall_adopted_similarity_stats
        }
        
        self.analysis_results = pd.DataFrame(results)
        return self.analysis_results
    
    def _calculate_stats(self, values: List[float]) -> Dict[str, float]:
        """数値リストの統計を計算"""
        if not values:
            return {'mean': None, 'max': None, 'min': None, 'std': None, 'count': 0}
        
        return {
            'mean': round(np.mean(values), 3),
            'max': round(np.max(values), 3),
            'min': round(np.min(values), 3),
            'std': round(np.std(values), 3),
            'count': len(values)
        }
    
    def _get_pattern_description(self, pattern: List[int], rag_columns: List[str]) -> str:
        """採択パターンの詳細説明を生成"""
        descriptions = []
        
        # 連続採択の分析
        consecutive_count = 0
        max_consecutive = 0
        for i, adopted in enumerate(pattern):
            if adopted:
                consecutive_count += 1
                max_consecutive = max(max_consecutive, consecutive_count)
            else:
                consecutive_count = 0
        
        if max_consecutive > 0:
            descriptions.append(f"最大連続採択: {max_consecutive}件")
        
        # 採択位置の分析
        adopted_positions = [i+1 for i, adopted in enumerate(pattern) if adopted]
        if adopted_positions:
            if len(adopted_positions) == 1:
                descriptions.append(f"採択位置: RAG{adopted_positions[0]}")
            else:
                descriptions.append(f"採択位置: RAG{adopted_positions[0]}-{adopted_positions[-1]}")
        
        return '; '.join(descriptions) if descriptions else '特になし'
    
    def analyze_adoption_patterns(self) -> Dict:
        """採択パターンの統計分析"""
        if self.analysis_results is None:
            raise ValueError("先にanalyze_rag_adoption_detailed()を実行してください。")
        
        patterns = self.analysis_results['採択パターン'].tolist()
        pattern_counts = Counter(patterns)
        
        # 最も一般的なパターン
        most_common_patterns = pattern_counts.most_common(5)
        
        # 採択数の分布
        adoption_counts = self.analysis_results['採択数'].tolist()
        count_distribution = Counter(adoption_counts)
        
        # 採択率の分布
        adoption_rates = self.analysis_results['採択率(%)'].tolist()
        
        analysis_summary = {
            '総行数': len(self.analysis_results),
            'ユニークパターン数': len(pattern_counts),
            '最も一般的なパターン': most_common_patterns,
            '採択数分布': dict(count_distribution),
            '平均採択率': np.mean(adoption_rates),
            '採択率の標準偏差': np.std(adoption_rates),
            '採択率の中央値': np.median(adoption_rates)
        }
        
        return analysis_summary
    
    def create_visualizations(self, output_dir: str = "analysis_plots"):
        """分析結果の可視化を作成"""
        if self.analysis_results is None:
            raise ValueError("先にanalyze_rag_adoption_detailed()を実行してください。")
        
        # 出力ディレクトリを作成
        os.makedirs(output_dir, exist_ok=True)
        
        # 日本語フォントの設定
        plt.rcParams['font.family'] = ['DejaVu Sans', 'Hiragino Sans', 'Yu Gothic', 'Meiryo', 'Takao', 'IPAexGothic', 'IPAPGothic', 'VL PGothic', 'Noto Sans CJK JP']
        
        # 1. 採択率の分布ヒストグラム
        plt.figure(figsize=(10, 6))
        plt.hist(self.analysis_results['採択率(%)'], bins=20, alpha=0.7, color='skyblue', edgecolor='black')
        plt.xlabel('採択率 (%)')
        plt.ylabel('行数')
        plt.title('RAG採択率の分布')
        plt.grid(True, alpha=0.3)
        plt.savefig(os.path.join(output_dir, 'adoption_rate_distribution.png'), dpi=300, bbox_inches='tight')
        plt.close()
        
        # 2. 採択数の分布
        plt.figure(figsize=(10, 6))
        adoption_counts = self.analysis_results['採択数'].value_counts().sort_index()
        plt.bar(adoption_counts.index, adoption_counts.values, alpha=0.7, color='lightcoral')
        plt.xlabel('採択数')
        plt.ylabel('行数')
        plt.title('RAG採択数の分布')
        plt.grid(True, alpha=0.3)
        plt.savefig(os.path.join(output_dir, 'adoption_count_distribution.png'), dpi=300, bbox_inches='tight')
        plt.close()
        
        # 3. Scoreの分布（採択・未採択別）
        if hasattr(self, 'overall_stats') and self.overall_stats['score']['count'] > 0:
            plt.figure(figsize=(12, 8))
            
            # 採択されたScoreと未採択のScoreを分けてプロット
            adopted_scores = self.analysis_results['採択Score平均'].dropna()
            all_scores = self.analysis_results['Score平均'].dropna()
            
            plt.subplot(2, 2, 1)
            plt.hist(all_scores, bins=20, alpha=0.7, color='lightblue', label='全体', edgecolor='black')
            plt.xlabel('Score平均')
            plt.ylabel('行数')
            plt.title('全体Score平均の分布')
            plt.grid(True, alpha=0.3)
            
            plt.subplot(2, 2, 2)
            if len(adopted_scores) > 0:
                plt.hist(adopted_scores, bins=20, alpha=0.7, color='gold', label='採択', edgecolor='black')
                plt.xlabel('採択Score平均')
                plt.ylabel('行数')
                plt.title('採択Score平均の分布')
                plt.grid(True, alpha=0.3)
            
            # Scoreと採択率の関係
            plt.subplot(2, 2, 3)
            valid_data = self.analysis_results[['採択率(%)', 'Score平均']].dropna()
            if len(valid_data) > 0:
                plt.scatter(valid_data['Score平均'], valid_data['採択率(%)'], alpha=0.6, color='blue')
                plt.xlabel('Score平均')
                plt.ylabel('採択率 (%)')
                plt.title('Score平均と採択率の関係')
                plt.grid(True, alpha=0.3)
            
            # Scoreの箱ひげ図
            plt.subplot(2, 2, 4)
            score_data = [all_scores]
            if len(adopted_scores) > 0:
                score_data.append(adopted_scores)
                labels = ['全体', '採択']
            else:
                labels = ['全体']
            
            plt.boxplot(score_data, labels=labels)
            plt.ylabel('Score')
            plt.title('Scoreの箱ひげ図')
            plt.grid(True, alpha=0.3)
            
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, 'score_analysis.png'), dpi=300, bbox_inches='tight')
            plt.close()
        
        # 4. 類似度の分布（採択・未採択別）
        if hasattr(self, 'overall_stats') and self.overall_stats['similarity']['count'] > 0:
            plt.figure(figsize=(12, 8))
            
            # 採択された類似度と未採択の類似度を分けてプロット
            adopted_similarities = self.analysis_results['採択類似度平均'].dropna()
            all_similarities = self.analysis_results['類似度平均'].dropna()
            
            plt.subplot(2, 2, 1)
            plt.hist(all_similarities, bins=20, alpha=0.7, color='lightgreen', label='全体', edgecolor='black')
            plt.xlabel('類似度平均')
            plt.ylabel('行数')
            plt.title('全体類似度平均の分布')
            plt.grid(True, alpha=0.3)
            
            plt.subplot(2, 2, 2)
            if len(adopted_similarities) > 0:
                plt.hist(adopted_similarities, bins=20, alpha=0.7, color='orange', label='採択', edgecolor='black')
                plt.xlabel('採択類似度平均')
                plt.ylabel('行数')
                plt.title('採択類似度平均の分布')
                plt.grid(True, alpha=0.3)
            
            # 類似度と採択率の関係
            plt.subplot(2, 2, 3)
            valid_data = self.analysis_results[['採択率(%)', '類似度平均']].dropna()
            if len(valid_data) > 0:
                plt.scatter(valid_data['類似度平均'], valid_data['採択率(%)'], alpha=0.6, color='green')
                plt.xlabel('類似度平均')
                plt.ylabel('採択率 (%)')
                plt.title('類似度平均と採択率の関係')
                plt.grid(True, alpha=0.3)
            
            # 類似度の箱ひげ図
            plt.subplot(2, 2, 4)
            similarity_data = [all_similarities]
            if len(adopted_similarities) > 0:
                similarity_data.append(adopted_similarities)
                labels = ['全体', '採択']
            else:
                labels = ['全体']
            
            plt.boxplot(similarity_data, labels=labels)
            plt.ylabel('類似度')
            plt.title('類似度の箱ひげ図')
            plt.grid(True, alpha=0.3)
            
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, 'similarity_analysis.png'), dpi=300, bbox_inches='tight')
            plt.close()
        
        # 5. 採択パターンのヒートマップ（上位10パターン）
        pattern_counts = Counter(self.analysis_results['採択パターン'])
        top_patterns = pattern_counts.most_common(10)
        
        if top_patterns:
            pattern_matrix = []
            pattern_labels = []
            
            for pattern, count in top_patterns:
                pattern_matrix.append([int(bit) for bit in pattern])
                pattern_labels.append(f"{pattern} ({count}回)")
            
            plt.figure(figsize=(12, 8))
            sns.heatmap(pattern_matrix, 
                       xticklabels=[f'RAG{i+1}' for i in range(len(pattern_matrix[0]))],
                       yticklabels=pattern_labels,
                       cmap='YlOrRd',
                       cbar_kws={'label': '採択状況 (1=採択, 0=未採択)'})
            plt.title('上位10パターンの採択状況ヒートマップ')
            plt.xlabel('RAGカラム')
            plt.ylabel('パターン (出現回数)')
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, 'adoption_patterns_heatmap.png'), dpi=300, bbox_inches='tight')
            plt.close()
        
        print(f"可視化ファイルを保存しました: {output_dir}")
    
    def generate_detailed_report(self, output_file: str = None):
        """詳細な分析レポートを生成"""
        if self.analysis_results is None:
            raise ValueError("先にanalyze_rag_adoption_detailed()を実行してください。")
        
        if output_file is None:
            base_name = os.path.splitext(self.excel_file_path)[0]
            output_file = f"{base_name}_detailed_analysis_report.xlsx"
        
        # パターン分析を実行
        pattern_analysis = self.analyze_adoption_patterns()
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 詳細分析結果
                self.analysis_results.to_excel(writer, sheet_name='詳細分析結果', index=False)
                
                # パターン分析サマリー
                pattern_summary = pd.DataFrame([
                    {'項目': k, '値': str(v)} for k, v in pattern_analysis.items()
                ])
                pattern_summary.to_excel(writer, sheet_name='パターン分析サマリー', index=False)
                
                # 全体統計サマリー
                if hasattr(self, 'overall_stats'):
                    overall_summary = []
                    for metric, stats in self.overall_stats.items():
                        if stats['count'] > 0:
                            overall_summary.append({
                                '指標': metric,
                                '平均値': stats['mean'],
                                '最大値': stats['max'],
                                '最小値': stats['min'],
                                '標準偏差': stats['std'],
                                'データ数': stats['count']
                            })
                    
                    if overall_summary:
                        overall_df = pd.DataFrame(overall_summary)
                        overall_df.to_excel(writer, sheet_name='Score・類似度統計', index=False)
                
                # 採択率ランキング
                ranking_df = self.analysis_results[['行番号', '採択率(%)', '採択数', '採択されたRAG']].sort_values('採択率(%)', ascending=False)
                ranking_df.to_excel(writer, sheet_name='採択率ランキング', index=False)
                
                # 高採択率の行（50%以上）
                high_adoption = self.analysis_results[self.analysis_results['採択率(%)'] >= 50]
                if len(high_adoption) > 0:
                    high_adoption.to_excel(writer, sheet_name='高採択率行', index=False)
                
                # 低採択率の行（10%以下）
                low_adoption = self.analysis_results[self.analysis_results['採択率(%)'] <= 10]
                if len(low_adoption) > 0:
                    low_adoption.to_excel(writer, sheet_name='低採択率行', index=False)
                
                # Score・類似度ランキング
                if hasattr(self, 'overall_stats') and self.overall_stats['score']['count'] > 0:
                    score_ranking = self.analysis_results[['行番号', 'Score平均', '採択Score平均', '採択率(%)']].dropna(subset=['Score平均']).sort_values('Score平均', ascending=False)
                    score_ranking.to_excel(writer, sheet_name='Scoreランキング', index=False)
                
                if hasattr(self, 'overall_stats') and self.overall_stats['similarity']['count'] > 0:
                    similarity_ranking = self.analysis_results[['行番号', '類似度平均', '採択類似度平均', '採択率(%)']].dropna(subset=['類似度平均']).sort_values('類似度平均', ascending=False)
                    similarity_ranking.to_excel(writer, sheet_name='類似度ランキング', index=False)
                
                # 列幅の自動調整
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"詳細分析レポートを保存しました: {output_file}")
            
        except Exception as e:
            print(f"レポート保存に失敗しました: {e}")
            raise
    
    def print_detailed_summary(self):
        """詳細なサマリーを表示"""
        if self.analysis_results is None:
            raise ValueError("先にanalyze_rag_adoption_detailed()を実行してください。")
        
        pattern_analysis = self.analyze_adoption_patterns()
        
        print("\n" + "="*60)
        print("詳細RAG分析結果サマリー")
        print("="*60)
        
        print(f"\n基本統計:")
        print(f"  総行数: {pattern_analysis['総行数']}")
        print(f"  ユニークパターン数: {pattern_analysis['ユニークパターン数']}")
        print(f"  平均採択率: {pattern_analysis['平均採択率']:.2f}%")
        print(f"  採択率の中央値: {pattern_analysis['採択率の中央値']:.2f}%")
        print(f"  採択率の標準偏差: {pattern_analysis['採択率の標準偏差']:.2f}%")
        
        # Scoreと類似度の統計を表示
        if hasattr(self, 'overall_stats'):
            print(f"\n=== Score統計 ===")
            score_stats = self.overall_stats['score']
            if score_stats['count'] > 0:
                print(f"  全体Score平均: {score_stats['mean']}")
                print(f"  全体Score最大: {score_stats['max']}")
                print(f"  全体Score最小: {score_stats['min']}")
                print(f"  全体Score標準偏差: {score_stats['std']}")
                print(f"  Scoreデータ数: {score_stats['count']}")
            
            adopted_score_stats = self.overall_stats['adopted_score']
            if adopted_score_stats['count'] > 0:
                print(f"  採択Score平均: {adopted_score_stats['mean']}")
                print(f"  採択Score最大: {adopted_score_stats['max']}")
                print(f"  採択Score最小: {adopted_score_stats['min']}")
                print(f"  採択Scoreデータ数: {adopted_score_stats['count']}")
            
            print(f"\n=== 類似度統計 ===")
            similarity_stats = self.overall_stats['similarity']
            if similarity_stats['count'] > 0:
                print(f"  全体類似度平均: {similarity_stats['mean']}")
                print(f"  全体類似度最大: {similarity_stats['max']}")
                print(f"  全体類似度最小: {similarity_stats['min']}")
                print(f"  全体類似度標準偏差: {similarity_stats['std']}")
                print(f"  類似度データ数: {similarity_stats['count']}")
            
            adopted_similarity_stats = self.overall_stats['adopted_similarity']
            if adopted_similarity_stats['count'] > 0:
                print(f"  採択類似度平均: {adopted_similarity_stats['mean']}")
                print(f"  採択類似度最大: {adopted_similarity_stats['max']}")
                print(f"  採択類似度最小: {adopted_similarity_stats['min']}")
                print(f"  採択類似度データ数: {adopted_similarity_stats['count']}")
        
        print(f"\n採択数分布:")
        for count, frequency in sorted(pattern_analysis['採択数分布'].items()):
            percentage = (frequency / pattern_analysis['総行数']) * 100
            print(f"  {count}件採択: {frequency}行 ({percentage:.1f}%)")
        
        print(f"\n最も一般的な採択パターン（上位5件）:")
        for i, (pattern, count) in enumerate(pattern_analysis['最も一般的なパターン'], 1):
            percentage = (count / pattern_analysis['総行数']) * 100
            adopted_count = pattern.count('1')
            print(f"  {i}. パターン {pattern}: {count}回 ({percentage:.1f}%) - {adopted_count}件採択")
        
        print(f"\n高採択率の行（採択率70%以上）:")
        high_adoption = self.analysis_results[self.analysis_results['採択率(%)'] >= 70]
        if len(high_adoption) > 0:
            for _, row in high_adoption.iterrows():
                print(f"  行{row['行番号']}: 採択率{row['採択率(%)']}% ({row['採択数']}/{row['総検索結果数']})")
        else:
            print("  採択率70%以上の行はありません")

def main():
    """メイン実行関数"""
    excel_file = "input.xlsx"
    
    # デバッグモード（色情報を表示するかどうか）
    debug_colors = False  # Trueにすると色情報が表示されます
    
    analyzer = RAGDetailedAnalysis(excel_file)
    
    try:
        print("RAG詳細分析を開始します...")
        
        # Excelファイルを読み込み
        analyzer.load_excel()
        
        # 詳細分析を実行
        result_df = analyzer.analyze_rag_adoption_detailed(debug_colors=debug_colors)
        
        # 詳細サマリーを表示
        analyzer.print_detailed_summary()
        
        # 可視化を作成
        analyzer.create_visualizations()
        
        # 詳細レポートを生成
        analyzer.generate_detailed_report()
        
        print("\n詳細分析が完了しました！")
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main() 