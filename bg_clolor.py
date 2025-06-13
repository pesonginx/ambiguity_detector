from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("sample.xlsx")
ws = wb["データシート"]

color_map = {}
color_index = 1

source_range = ws["A1:C10"]
column_offset = 3

for row in source_range:
    for cell in row:
        fill = cell.fill

        # 色の識別キーを生成
        if fill and isinstance(fill, PatternFill):
            fg = fill.fgColor

            # 色タイプ別に一意なキーを生成
            if fg.type == "rgb":
                color_key = f"rgb:{fg.rgb}"
            elif fg.type == "theme":
                color_key = f"theme:{fg.theme}"
            elif fg.type == "indexed":
                color_key = f"idx:{fg.indexed}"
            else:
                color_key = "NO_COLOR"
        else:
            color_key = "NO_COLOR"

        # 初めての色なら番号を割り当て
        if color_key not in color_map:
            color_map[color_key] = color_index
            color_index += 1

        # 対応する右のセルに番号を出力
        ws.cell(row=cell.row, column=cell.column + column_offset).value = color_map[color_key]

wb.save("output_colored.xlsx")
