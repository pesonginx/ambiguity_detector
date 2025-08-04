from openpyxl import load_workbook
import pandas as pd

def extract_used_docs_from_colored_cells(filename, id_column="query_id"):
    wb = load_workbook(filename)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    results = []

    for row in sheet.iter_rows(min_row=2):
        query_id = row[0].value
        used_docs = []
        for cell, header in zip(row[1:], headers[1:]):
            if cell.fill and cell.fill.fill_type == "solid":
                fg = cell.fill.fgColor.rgb
                if fg and fg not in ["FFFFFFFF", "00000000"]:  # 白 or 無色でなければ色付きとみなす
                    used_docs.append(header)
        results.append({"query_id": query_id, "used_docs": used_docs})
    return pd.DataFrame(results)





import pandas as pd
from scipy.stats import ttest_rel, wilcoxon, chi2_contingency
import matplotlib.pyplot as plt

# --- ファイル読み込み ---
df_a = pd.read_excel("rag_a.xlsx")  # query_id, topk_docs
df_b = pd.read_excel("rag_b.xlsx")
df_llm = extract_used_docs_from_colored_cells("llm_used.xlsx")

# --- 前処理 ---
# カンマ区切りの文字列 → リスト
for df in [df_a, df_b, df_llm]:
    df.columns = [col.lower() for col in df.columns]  # 小文字化（安全のため）
    df["topk_docs" if "topk_docs" in df.columns else "used_docs"] = \
        df["topk_docs" if "topk_docs" in df.columns else "used_docs"].fillna("").apply(lambda x: x.split(",") if x else [])

# --- 結合 ---
merged = df_llm.merge(df_a, on="query_id", suffixes=("", "_a")).merge(
    df_b, on="query_id", suffixes=("", "_b"))
merged = merged.rename(columns={"topk_docs": "rag_a_topk", "topk_docs_b": "rag_b_topk", "used_docs": "llm_used_docs"})

# --- 採択率計算 ---
merged["a_hit"] = merged.apply(lambda row: len(set(row["rag_a_topk"]) & set(row["llm_used_docs"])), axis=1)
merged["b_hit"] = merged.apply(lambda row: len(set(row["rag_b_topk"]) & set(row["llm_used_docs"])), axis=1)
merged["a_hit_ratio"] = merged["a_hit"] / merged["rag_a_topk"].apply(len)
merged["b_hit_ratio"] = merged["b_hit"] / merged["rag_b_topk"].apply(len)

# --- 統計分析（t検定） ---
t_stat, p_val = ttest_rel(merged["a_hit_ratio"], merged["b_hit_ratio"])
print(f"[t検定] 採択率比較: t={t_stat:.3f}, p={p_val:.3f}")

# --- Jaccard類似度 ---
def jaccard(a, b):
    return len(set(a) & set(b)) / len(set(a) | set(b)) if (a or b) else 0
merged["jaccard_topk"] = merged.apply(lambda row: jaccard(row["rag_a_topk"], row["rag_b_topk"]), axis=1)

# --- 採択元カテゴリ ---
def adoption_source(row):
    in_a = any(doc in row["rag_a_topk"] for doc in row["llm_used_docs"])
    in_b = any(doc in row["rag_b_topk"] for doc in row["llm_used_docs"])
    if in_a and in_b:
        return "both"
    elif in_a:
        return "only_a"
    elif in_b:
        return "only_b"
    else:
        return "neither"
merged["adopt_type"] = merged.apply(adoption_source, axis=1)

# --- カイ二乗検定 ---
adopt_counts = merged["adopt_type"].value_counts().reindex(["only_a", "only_b", "both", "neither"], fill_value=0)
chi2, p, dof, expected = chi2_contingency([adopt_counts.values])
print(f"[カイ二乗検定] 採択元の偏り: chi2={chi2:.3f}, p={p:.3f}")

# --- 可視化 ---
plt.boxplot([merged["a_hit_ratio"], merged["b_hit_ratio"]], labels=["RAG-A", "RAG-B"])
plt.title("LLM採択率比較")
plt.ylabel("採択率")
plt.grid(True)
plt.show()

# --- 結果保存 ---
merged.to_excel("merged_rag_analysis.xlsx", index=False)
