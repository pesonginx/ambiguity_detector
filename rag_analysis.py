import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from typing import List, Tuple, Dict
import os
import re

class RAGAnalysis:
    """RAG検索結果の分析クラス"""
    
    def __init__(self, excel_file_path: str):
        """
        RAG分析クラスの初期化
        
        Parameters:
        -----------
        excel_file_path : str
            分析対象のExcelファイルパス
        """
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.worksheet = None
        self.df = None
        
    def load_excel(self, sheet_name: str = None):
        """
        Excelファイルを読み込み
        
        Parameters:
        -----------
        sheet_name : str, optional
            読み込むシート名（Noneの場合は最初のシート）
        """
        try:
            # pandasでデータを読み込み
            excel_data = pd.read_excel(self.excel_file_path, sheet_name=sheet_name)
            
            # 複数シートの場合は辞書形式で返されるため、適切に処理
            if isinstance(excel_data, dict):
                print(f"複数シートが検出されました。利用可能なシート: {list(excel_data.keys())}")
                if sheet_name is None:
                    # シート名が指定されていない場合は最初のシートを使用
                    sheet_name = list(excel_data.keys())[0]
                    print(f"最初のシート '{sheet_name}' を使用します。")
                elif sheet_name not in excel_data:
                    print(f"指定されたシート '{sheet_name}' が見つかりません。最初のシートを使用します。")
                    sheet_name = list(excel_data.keys())[0]
                self.df = excel_data[sheet_name]
            else:
                self.df = excel_data
            
            # openpyxlでワークブックを読み込み（色情報取得用）
            self.workbook = openpyxl.load_workbook(self.excel_file_path)
            
            # シート名の処理
            if sheet_name is None:
                self.worksheet = self.workbook.active
            else:
                try:
                    self.worksheet = self.workbook[sheet_name]
                except KeyError:
                    print(f"警告: シート '{sheet_name}' が見つかりません。最初のシートを使用します。")
                    self.worksheet = self.workbook.active
            
            print(f"Excelファイルを読み込みました: {self.excel_file_path}")
            print(f"使用シート: {sheet_name if sheet_name else '最初のシート'}")
            print(f"データ形状: {self.df.shape}")
            print(f"カラム: {list(self.df.columns)}")
            
            # Excelファイルの基本情報を表示
            print(f"ワークブックのシート数: {len(self.workbook.sheetnames)}")
            print(f"利用可能なシート: {self.workbook.sheetnames}")
            print(f"現在のワークシート: {self.worksheet.title}")
            print(f"ワークシートの範囲: A1:{self.worksheet.max_column}{self.worksheet.max_row}")
            
        except Exception as e:
            print(f"Excelファイルの読み込みに失敗しました: {e}")
            raise
    
    def is_colored_cell(self, cell) -> bool:
        """
        セルが塗りつぶされているかチェック（白または無色以外）
        
        Parameters:
        -----------
        cell : openpyxl.cell.Cell
            チェック対象のセル
            
        Returns:
        --------
        bool
            塗りつぶされている場合True
        """
        # 複数の方法で色情報を取得
        fill_objects = []
        
        # 方法1: 直接のfill属性
        if hasattr(cell, 'fill') and cell.fill:
            fill_objects.append(cell.fill)
        
        # 方法2: style.fill属性
        if hasattr(cell, 'style') and hasattr(cell.style, 'fill') and cell.style.fill:
            fill_objects.append(cell.style.fill)
        
        # 方法3: StyleProxyの内部オブジェクト
        for fill in fill_objects[:]:  # コピーを作成してループ中に変更
            if hasattr(fill, '_style') and hasattr(fill._style, 'fill'):
                fill_objects.append(fill._style.fill)
            elif hasattr(fill, 'fill'):
                fill_objects.append(fill.fill)
        
        # 各fillオブジェクトをチェック
        for fill in fill_objects:
            if not fill:
                continue
                
            if not isinstance(fill, PatternFill):
                continue
                
            fg_color = fill.fgColor
            bg_color = fill.bgColor
            pattern_type = fill.patternType
            
            # パターンが設定されている場合
            if pattern_type and pattern_type != 'none':
                return True
            
            # 前景色の判定
            if fg_color and fg_color.type != "auto":
                if self._is_colored_color(fg_color):
                    return True
            
            # 背景色の判定
            if bg_color and bg_color.type != "auto":
                if self._is_colored_color(bg_color):
                    return True
                    
        return False
    
    def _is_colored_color(self, color) -> bool:
        """Colorオブジェクトが白以外の色かどうかを判定"""
        if not color:
            return False
            
        if color.type == "rgb":
            # RGB値で白を判定
            if color.rgb and color.rgb.upper() in ["FFFFFF", "FFFFFFFF", "00000000"]:
                return False
            # RGB値が設定されている場合は塗りつぶされている
            if color.rgb:
                return True
        elif color.type == "theme":
            # テーマ色で白を判定（一般的な白のテーマ番号）
            if color.theme in [0, 1]:  # 白系のテーマ番号
                return False
            # テーマ色が設定されている場合は塗りつぶされている
            if color.theme is not None:
                return True
        elif color.type == "indexed":
            # インデックス色で白を判定
            if color.indexed in [0, 1]:  # 白系のインデックス番号
                return False
            # インデックス色が設定されている場合は塗りつぶされている
            if color.indexed is not None:
                return True
                
        return False
    
    def check_conditional_formatting(self, cell) -> bool:
        """
        条件付き書式で色が設定されているかチェック
        
        Parameters:
        -----------
        cell : openpyxl.cell.Cell
            チェック対象のセル
            
        Returns:
        --------
        bool
            条件付き書式で色が設定されている場合True
        """
        # 条件付き書式をチェック
        for cf in self.worksheet.conditional_formatting:
            for rule in cf.cf_rules:
                if hasattr(rule, 'dxf') and rule.dxf and hasattr(rule.dxf, 'fill'):
                    # 条件付き書式の範囲をチェック
                    for range_str in cf.sqref.ranges:
                        if cell.coordinate in range_str:
                            return True
        return False
    
    def debug_cell_color(self, cell) -> str:
        """
        セルの色情報をデバッグ用に取得
        
        Parameters:
        -----------
        cell : openpyxl.cell.Cell
            チェック対象のセル
            
        Returns:
        --------
        str
            色情報の文字列
        """
        # 複数の方法で色情報を取得
        fill_objects = []
        
        # 方法1: 直接のfill属性
        if hasattr(cell, 'fill') and cell.fill:
            fill_objects.append(cell.fill)
        
        # 方法2: style.fill属性
        if hasattr(cell, 'style') and hasattr(cell.style, 'fill') and cell.style.fill:
            fill_objects.append(cell.style.fill)
        
        # 方法3: StyleProxyの内部オブジェクト
        for fill in fill_objects[:]:  # コピーを作成してループ中に変更
            if hasattr(fill, '_style') and hasattr(fill._style, 'fill'):
                fill_objects.append(fill._style.fill)
            elif hasattr(fill, 'fill'):
                fill_objects.append(fill.fill)
        
        if not fill_objects:
            return "No fill objects found"
        
        color_info = []
        for i, fill in enumerate(fill_objects):
            if not fill:
                color_info.append(f"Fill{i}: None")
                continue
                
            if not isinstance(fill, PatternFill):
                color_info.append(f"Fill{i}: {type(fill)}")
                continue
                
            fg_color = fill.fgColor
            bg_color = fill.bgColor
            pattern_type = fill.patternType
            
            fill_info = f"Fill{i}: Pattern={pattern_type}"
            
            # 前景色の詳細情報
            if fg_color:
                fill_info += f", FG={self._get_color_info(fg_color)}"
            else:
                fill_info += ", FG=None"
                
            # 背景色の詳細情報
            if bg_color:
                fill_info += f", BG={self._get_color_info(bg_color)}"
            else:
                fill_info += ", BG=None"
                
            color_info.append(fill_info)
            
        return " | ".join(color_info)
    
    def _get_color_info(self, color) -> str:
        """Colorオブジェクトから色情報を取得"""
        if not color:
            return "None"
            
        color_info = f"{color.type}"
        
        if color.type == "rgb":
            color_info += f"={color.rgb}"
        elif color.type == "theme":
            color_info += f"={color.theme}"
        elif color.type == "indexed":
            color_info += f"={color.indexed}"
        elif color.type == "auto":
            color_info += "=auto"
        else:
            color_info += f"=unknown({color})"
            
        return color_info
    
    def get_rag_columns(self) -> List[str]:
        """
        RAGカラム（RAG1〜RAG10）を取得
        
        Returns:
        --------
        List[str]
            RAGカラム名のリスト
        """
        rag_columns = []
        for i in range(1, 11):
            column_name = f"RAG{i}"
            if column_name in self.df.columns:
                rag_columns.append(column_name)
        
        return rag_columns
    
    def extract_score_and_similarity(self, text: str) -> Dict[str, float]:
        """
        RAGカラムのテキストからScoreと類似度を抽出
        
        Parameters:
        -----------
        text : str
            抽出対象のテキスト
            
        Returns:
        --------
        Dict[str, float]
            Scoreと類似度の値（抽出できない場合はNone）
        """
        if pd.isna(text) or not isinstance(text, str):
            return {'score': None, 'similarity': None}
        
        result = {'score': None, 'similarity': None}
        
        # Scoreの抽出パターン
        score_patterns = [
            r'Score[:\s]*([0-9]+\.?[0-9]*)',
            r'スコア[:\s]*([0-9]+\.?[0-9]*)',
            r'検索スコア[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*\(Score\)',
            r'Score\s*=\s*([0-9]+\.?[0-9]*)'
        ]
        
        # 類似度の抽出パターン
        similarity_patterns = [
            r'類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'LLM回答\s*⇔\s*【回答】類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'回答類似度[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*\(類似度\)',
            r'Similarity[:\s]*([0-9]+\.?[0-9]*)',
            r'([0-9]+\.?[0-9]*)\s*%?\s*\(類似度\)'
        ]
        
        # Scoreの抽出
        for pattern in score_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    result['score'] = float(match.group(1))
                    break
                except ValueError:
                    continue
        
        # 類似度の抽出
        for pattern in similarity_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    result['similarity'] = float(match.group(1))
                    break
                except ValueError:
                    continue
        
        return result
    
    def analyze_rag_adoption(self, debug_colors: bool = False) -> pd.DataFrame:
        """
        RAG検索結果の採択分析を実行
        
        Parameters:
        -----------
        debug_colors : bool, optional
            色情報のデバッグ出力を行うかどうか（デフォルト: False）
            
        Returns:
        --------
        pd.DataFrame
            採択分析結果を含むDataFrame
        """
        if self.df is None:
            raise ValueError("Excelファイルが読み込まれていません。load_excel()を先に実行してください。")
        
        rag_columns = self.get_rag_columns()
        if not rag_columns:
            raise ValueError("RAGカラム（RAG1〜RAG10）が見つかりません。")
        
        print(f"分析対象RAGカラム: {rag_columns}")
        
        # 結果を格納するリスト
        results = []
        
        # 全体のScoreと類似度の統計用
        all_scores = []
        all_similarities = []
        adopted_scores = []
        adopted_similarities = []
        
        # 各行を分析
        for row_idx, row_data in self.df.iterrows():
            adopted_count = 0
            adopted_rags = []
            row_scores = []
            row_similarities = []
            adopted_row_scores = []
            adopted_row_similarities = []
            
            # 各RAGカラムをチェック
            for rag_col in rag_columns:
                # Excelのセル座標を取得（1ベース）
                excel_row = row_idx + 2  # ヘッダー行を考慮
                excel_col = self.df.columns.get_loc(rag_col) + 1
                
                # デバッグ出力（座標確認）
                if debug_colors and row_idx < 2:
                    print(f"  座標: 行{excel_row}, 列{excel_col} ({rag_col})")
                
                # セルを取得
                cell = self.worksheet.cell(row=excel_row, column=excel_col)
                
                # セルが塗りつぶされているかチェック
                is_adopted = self.is_colored_cell(cell)
                
                # 条件付き書式もチェック
                if not is_adopted:
                    is_adopted = self.check_conditional_formatting(cell)
                
                # デバッグ出力（必要に応じて）
                if debug_colors and row_idx < 5:  # 最初の5行のみデバッグ出力
                    color_info = self.debug_cell_color(cell)
                    print(f"行{row_idx+1}, {rag_col}: {color_info} -> {'採択' if is_adopted else '未採択'}")
                
                # Scoreと類似度を抽出
                cell_text = str(row_data[rag_col]) if not pd.isna(row_data[rag_col]) else ""
                extracted_values = self.extract_score_and_similarity(cell_text)
                
                # 統計用に値を記録
                if extracted_values['score'] is not None:
                    row_scores.append(extracted_values['score'])
                    all_scores.append(extracted_values['score'])
                    if is_adopted:
                        adopted_row_scores.append(extracted_values['score'])
                        adopted_scores.append(extracted_values['score'])
                
                if extracted_values['similarity'] is not None:
                    row_similarities.append(extracted_values['similarity'])
                    all_similarities.append(extracted_values['similarity'])
                    if is_adopted:
                        adopted_row_similarities.append(extracted_values['similarity'])
                        adopted_similarities.append(extracted_values['similarity'])
                
                if is_adopted:
                    adopted_count += 1
                    adopted_rags.append(rag_col)
            
            # 採択率を計算
            adoption_rate = (adopted_count / len(rag_columns)) * 100 if rag_columns else 0
            
            # 行内のScoreと類似度の統計
            row_score_stats = self._calculate_stats(row_scores)
            row_similarity_stats = self._calculate_stats(row_similarities)
            adopted_score_stats = self._calculate_stats(adopted_row_scores)
            adopted_similarity_stats = self._calculate_stats(adopted_row_similarities)
            
            # 結果を記録
            result = {
                '行番号': row_idx + 1,
                '採択数': adopted_count,
                '総検索結果数': len(rag_columns),
                '採択率(%)': round(adoption_rate, 2),
                '採択されたRAG': ', '.join(adopted_rags) if adopted_rags else 'なし',
                # Score統計
                'Score平均': row_score_stats['mean'],
                'Score最大': row_score_stats['max'],
                'Score最小': row_score_stats['min'],
                'Score標準偏差': row_score_stats['std'],
                '採択Score平均': adopted_score_stats['mean'],
                '採択Score最大': adopted_score_stats['max'],
                '採択Score最小': adopted_score_stats['min'],
                # 類似度統計
                '類似度平均': row_similarity_stats['mean'],
                '類似度最大': row_similarity_stats['max'],
                '類似度最小': row_similarity_stats['min'],
                '類似度標準偏差': row_similarity_stats['std'],
                '採択類似度平均': adopted_similarity_stats['mean'],
                '採択類似度最大': adopted_similarity_stats['max'],
                '採択類似度最小': adopted_similarity_stats['min']
            }
            
            # 元のデータも含める
            for col in self.df.columns:
                result[col] = row_data[col]
            
            results.append(result)
        
        # 全体統計を計算
        overall_score_stats = self._calculate_stats(all_scores)
        overall_similarity_stats = self._calculate_stats(all_similarities)
        overall_adopted_score_stats = self._calculate_stats(adopted_scores)
        overall_adopted_similarity_stats = self._calculate_stats(adopted_similarities)
        
        # 全体統計を保存
        self.overall_stats = {
            'score': overall_score_stats,
            'similarity': overall_similarity_stats,
            'adopted_score': overall_adopted_score_stats,
            'adopted_similarity': overall_adopted_similarity_stats
        }
        
        # DataFrameに変換
        result_df = pd.DataFrame(results)
        
        return result_df
    
    def _calculate_stats(self, values: List[float]) -> Dict[str, float]:
        """数値リストの統計を計算"""
        if not values:
            return {'mean': None, 'max': None, 'min': None, 'std': None, 'count': 0}
        
        return {
            'mean': round(np.mean(values), 3),
            'max': round(np.max(values), 3),
            'min': round(np.min(values), 3),
            'std': round(np.std(values), 3),
            'count': len(values)
        }
    
    def save_analysis_result(self, result_df: pd.DataFrame, output_file: str = None):
        """
        分析結果をExcelファイルに保存
        
        Parameters:
        -----------
        result_df : pd.DataFrame
            保存する分析結果
        output_file : str, optional
            出力ファイル名（Noneの場合は自動生成）
        """
        if output_file is None:
            base_name = os.path.splitext(self.excel_file_path)[0]
            output_file = f"{base_name}_analysis_result.xlsx"
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 詳細分析結果
                result_df.to_excel(writer, sheet_name='RAG分析結果', index=False)
                
                # 全体統計サマリー
                if hasattr(self, 'overall_stats'):
                    overall_summary = []
                    for metric, stats in self.overall_stats.items():
                        if stats['count'] > 0:
                            overall_summary.append({
                                '指標': metric,
                                '平均値': stats['mean'],
                                '最大値': stats['max'],
                                '最小値': stats['min'],
                                '標準偏差': stats['std'],
                                'データ数': stats['count']
                            })
                    
                    if overall_summary:
                        overall_df = pd.DataFrame(overall_summary)
                        overall_df.to_excel(writer, sheet_name='全体統計サマリー', index=False)
                
                # 列幅の自動調整
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"分析結果を保存しました: {output_file}")
            
        except Exception as e:
            print(f"ファイル保存に失敗しました: {e}")
            raise
    
    def print_summary(self, result_df: pd.DataFrame):
        """
        分析結果のサマリーを表示
        
        Parameters:
        -----------
        result_df : pd.DataFrame
            分析結果のDataFrame
        """
        print("\n=== RAG分析結果サマリー ===")
        print(f"総行数: {len(result_df)}")
        print(f"平均採択率: {result_df['採択率(%)'].mean():.2f}%")
        print(f"最大採択率: {result_df['採択率(%)'].max():.2f}%")
        print(f"最小採択率: {result_df['採択率(%)'].min():.2f}%")
        print(f"標準偏差: {result_df['採択率(%)'].std():.2f}%")
        
        # Scoreと類似度の統計を表示
        if hasattr(self, 'overall_stats'):
            print("\n=== Score統計 ===")
            score_stats = self.overall_stats['score']
            if score_stats['count'] > 0:
                print(f"全体Score平均: {score_stats['mean']}")
                print(f"全体Score最大: {score_stats['max']}")
                print(f"全体Score最小: {score_stats['min']}")
                print(f"全体Score標準偏差: {score_stats['std']}")
                print(f"Scoreデータ数: {score_stats['count']}")
            
            adopted_score_stats = self.overall_stats['adopted_score']
            if adopted_score_stats['count'] > 0:
                print(f"採択Score平均: {adopted_score_stats['mean']}")
                print(f"採択Score最大: {adopted_score_stats['max']}")
                print(f"採択Score最小: {adopted_score_stats['min']}")
                print(f"採択Scoreデータ数: {adopted_score_stats['count']}")
            
            print("\n=== 類似度統計 ===")
            similarity_stats = self.overall_stats['similarity']
            if similarity_stats['count'] > 0:
                print(f"全体類似度平均: {similarity_stats['mean']}")
                print(f"全体類似度最大: {similarity_stats['max']}")
                print(f"全体類似度最小: {similarity_stats['min']}")
                print(f"全体類似度標準偏差: {similarity_stats['std']}")
                print(f"類似度データ数: {similarity_stats['count']}")
            
            adopted_similarity_stats = self.overall_stats['adopted_similarity']
            if adopted_similarity_stats['count'] > 0:
                print(f"採択類似度平均: {adopted_similarity_stats['mean']}")
                print(f"採択類似度最大: {adopted_similarity_stats['max']}")
                print(f"採択類似度最小: {adopted_similarity_stats['min']}")
                print(f"採択類似度データ数: {adopted_similarity_stats['count']}")
        
        print("\n=== 採択率分布 ===")
        adoption_ranges = [
            (0, 10, "0-10%"),
            (10, 30, "10-30%"),
            (30, 50, "30-50%"),
            (50, 70, "50-70%"),
            (70, 90, "70-90%"),
            (90, 101, "90-100%")
        ]
        
        for min_rate, max_rate, label in adoption_ranges:
            count = len(result_df[(result_df['採択率(%)'] >= min_rate) & (result_df['採択率(%)'] < max_rate)])
            percentage = (count / len(result_df)) * 100
            print(f"{label}: {count}行 ({percentage:.1f}%)")
        
        print("\n=== 高採択率の行（採択率50%以上） ===")
        high_adoption = result_df[result_df['採択率(%)'] >= 50]
        if len(high_adoption) > 0:
            for _, row in high_adoption.iterrows():
                print(f"行{row['行番号']}: 採択率{row['採択率(%)']}% ({row['採択数']}/{row['総検索結果数']})")
        else:
            print("採択率50%以上の行はありません")

def main():
    """メイン実行関数"""
    # 分析対象のExcelファイル
    excel_file = "input.xlsx"
    
    # デバッグモード（色情報を表示するかどうか）
    debug_colors = True  # Trueにすると色情報が表示されます
    
    # RAG分析クラスのインスタンスを作成
    analyzer = RAGAnalysis(excel_file)
    
    try:
        # Excelファイルを読み込み
        analyzer.load_excel()
        
        # RAG採択分析を実行
        result_df = analyzer.analyze_rag_adoption(debug_colors=debug_colors)
        
        # サマリーを表示
        analyzer.print_summary(result_df)
        
        # 結果を保存
        analyzer.save_analysis_result(result_df)
        
        print("\n分析が完了しました！")
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main() 